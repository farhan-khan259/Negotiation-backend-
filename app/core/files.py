import os
import shutil
from pathlib import Path
from typing import Optional, List
from fastapi import UploadFile
import PyPDF2
from docx import Document
from PIL import Image
from pdf2image import convert_from_path
import pytesseract
from openpyxl import load_workbook
from supabase import create_client, Client
from app.core.config import settings

UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

MIN_TEXT_LENGTH_FOR_OCR = 50


def _configure_tesseract() -> None:
    """Configure Tesseract binary path if provided via env."""
    tesseract_cmd = os.getenv("OCR_TESSERACT_CMD")
    if tesseract_cmd:
        pytesseract.pytesseract.tesseract_cmd = tesseract_cmd

def get_supabase_client() -> Optional[Client]:
    if settings.SUPABASE_URL and settings.SUPABASE_KEY:
        return create_client(settings.SUPABASE_URL, settings.SUPABASE_KEY)
    return None

async def upload_to_supabase(file_path: str, file_name: str, content_type: str) -> Optional[str]:
    """
    Uploads a file to Supabase Storage and returns the public URL.
    """
    supabase = get_supabase_client()
    if not supabase:
        return None

    try:
        bucket_name = settings.SUPABASE_BUCKET_NAME
        # Ensure bucket exists (optional, usually created manually or via migration)
        # supabase.storage.create_bucket(bucket_name) 

        with open(file_path, 'rb') as f:
            file_bytes = f.read()
        
        # Upload
        storage_path = f"{file_name}"
        supabase.storage.from_(bucket_name).upload(
            path=storage_path,
            file=file_bytes,
            file_options={"content-type": content_type, "upsert": "true"}
        )
        
        # Get Public URL
        public_url = supabase.storage.from_(bucket_name).get_public_url(storage_path)
        return public_url
    except Exception as e:
        print(f"Error uploading to Supabase: {e}")
        return None

async def save_upload_file(upload_file: UploadFile) -> str:
    """
    Saves an uploaded file to the local disk and returns the file path.
    """
    try:
        file_path = UPLOAD_DIR / upload_file.filename
        with file_path.open("wb") as buffer:
            shutil.copyfileobj(upload_file.file, buffer)
        return str(file_path)
    finally:
        upload_file.file.close()

def extract_text_from_pdf(file_path: str) -> Optional[str]:
    """
    Extracts text from a PDF file.
    """
    try:
        text = ""
        with open(file_path, "rb") as file:
            reader = PyPDF2.PdfReader(file)
            for page in reader.pages:
                text += page.extract_text() + "\n"
        return text
    except Exception as e:
        print(f"Error extracting text from PDF: {e}")
        return None


def extract_text_from_image(file_path: str) -> Optional[str]:
    """
    Extract text from an image using OCR.
    """
    try:
        _configure_tesseract()
        with Image.open(file_path) as img:
            return pytesseract.image_to_string(img)
    except Exception as e:
        print(f"Error extracting text from image: {e}")
        return None


def extract_text_from_pdf_ocr(file_path: str) -> Optional[str]:
    """
    Extract text from a scanned PDF using OCR.
    """
    try:
        _configure_tesseract()
        pages = convert_from_path(file_path)
        texts: List[str] = []
        for page in pages:
            texts.append(pytesseract.image_to_string(page))
        return "\n".join([t for t in texts if t])
    except Exception as e:
        print(f"Error extracting text from PDF via OCR: {e}")
        return None

def extract_text_from_docx(file_path: str) -> Optional[str]:
    """
    Extract text from DOCX file.
    """
    try:
        doc = Document(file_path)
        full_text = []
        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                full_text.append(paragraph.text)
        return "\n".join(full_text)
    except Exception as e:
        print(f"Error extracting text from DOCX: {e}")
        return None


def extract_text_from_excel(file_path: str) -> Optional[str]:
    """
    Extract text from Excel file (.xlsx, .xls).
    Reads all sheets and cells, formatting as readable text.
    """
    try:
        workbook = load_workbook(file_path, data_only=True, read_only=True)
        full_text = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            full_text.append(f"Sheet: {sheet_name}")
            full_text.append("=" * 50)
            
            # Read rows and format as text
            for row in sheet.iter_rows(values_only=True):
                # Filter out None values and convert to strings
                row_values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if row_values:
                    full_text.append(" | ".join(row_values))
            
            full_text.append("")  # Empty line between sheets
        
        return "\n".join(full_text)
    except Exception as e:
        print(f"Error extracting text from Excel: {e}")
        return None


def extract_text_from_file(file_path: str, content_type: str) -> Optional[str]:
    """
    Extracts text based on file type.
    """
    print(f"DEBUG: Extracting text from {file_path} (type: {content_type})")
    
    if content_type == "application/pdf":
        text = extract_text_from_pdf(file_path)
        if not text or len(text.strip()) < MIN_TEXT_LENGTH_FOR_OCR:
            print("INFO: Low text detected in PDF, attempting OCR...")
            ocr_text = extract_text_from_pdf_ocr(file_path)
            if ocr_text and len(ocr_text.strip()) > len(text or ""):
                text = ocr_text
    elif content_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        text = extract_text_from_docx(file_path)
    elif content_type in [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
        "application/vnd.ms-excel"  # .xls
    ]:
        text = extract_text_from_excel(file_path)
    elif content_type.startswith("image/"):
        text = extract_text_from_image(file_path)
    elif content_type.startswith("text/"):
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                text = f.read()
        except Exception as e:
            print(f"Error reading text file: {e}")
            return None
    else:
        print(f"WARNING: Unsupported file type: {content_type}")
        return None
    
    if text:
        print(f"DEBUG: Extracted {len(text)} characters from file")
    else:
        print(f"WARNING: No text extracted from file")
    
    return text
